#Import file to read write data from excel
import openpyxl as xl
#Import to add chart to a file
from openpyxl.chart import BarChart,Reference

#Main function
def process_workbook(filename):

    #Loading the excel file
    wb = xl.load_workbook(filename)
    Sheet = wb['Sheet1']

    #Code for updating price in each cell
    for row in range(2,Sheet.max_row+1):
        cell = Sheet.cell(row,3)
        corrected_price = cell.value*0.9
    #Saving the updated prices in new column
        corrected_price_cell = Sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    #Fetching values for chart
    values = Reference(Sheet, min_row=2, max_row=Sheet.max_row, min_col=4, max_col=4)

    chart = BarChart
    chart.add_data
    Sheet.add_chart(chart,'e2')

    wb.save(filename)